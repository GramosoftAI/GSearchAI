import io
import openpyxl
from typing import Dict, Any
from ..domain.pipeline_context import PipelineContext
from ..domain.workbook import Workbook, Sheet, LogicalTable

class WorkbookDiscoveryEngine:
    """Discovers raw sheets, reads formats, and detects tables via openpyxl."""
    def run(self, context: PipelineContext) -> PipelineContext:
        ext = context.filename.lower().split(".")[-1] if "." in context.filename else ""
        if ext == "csv":
            try:
                import csv
                text_content = context.file_bytes.decode('utf-8', errors='replace')
                reader = csv.DictReader(io.StringIO(text_content))
                raw_data = list(reader)
                
                workbook_domain = Workbook(
                    filename=context.filename,
                    file_size_bytes=len(context.file_bytes),
                    sheets=[]
                )
                
                if raw_data:
                    table = LogicalTable(
                        table_id="csv_tbl1",
                        start_row=1,
                        end_row=len(raw_data),
                        raw_data=raw_data
                    )
                    sheet = Sheet(name="Sheet1", classification="Raw Data", tables=[table])
                    workbook_domain.sheets.append(sheet)
                    
                context.workbook = workbook_domain
                return context
            except Exception as e:
                context.add_error(f"CSV Parsing failed: {e}")
                return context
                
        if ext not in ["xlsx", "xlsm"]:
            context.add_warning("Not an openpyxl-compatible format. Fallback to basic processing needed.")
            return context
            
        try:
            wb = openpyxl.load_workbook(io.BytesIO(context.file_bytes), data_only=True)
            workbook_domain = Workbook(
                filename=context.filename,
                file_size_bytes=len(context.file_bytes),
                sheets=[]
            )
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Basic LogicalTable extraction (whole sheet for now)
                raw_data = []
                headers = []
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if row_idx == 0:
                        headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(row)]
                        continue
                        
                    row_dict = {}
                    has_data = False
                    for i, val in enumerate(row):
                        if val is not None and str(val).strip() != "":
                            has_data = True
                        row_dict[headers[i] if i < len(headers) else f"col_{i}"] = val
                    
                    if has_data:
                        raw_data.append(row_dict)
                        
                if raw_data:
                    table = LogicalTable(
                        table_id=f"{sheet_name}_tbl1",
                        start_row=1,
                        end_row=len(raw_data),
                        raw_data=raw_data
                    )
                    sheet = Sheet(name=sheet_name, classification="Raw Data", tables=[table])
                    workbook_domain.sheets.append(sheet)
                    
            context.workbook = workbook_domain
            
        except Exception as e:
            context.add_error(f"WorkbookDiscoveryEngine failed: {e}")
            
        return context
